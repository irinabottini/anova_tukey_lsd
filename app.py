from __future__ import annotations

import io
import threading
import time
from itertools import combinations
from typing import Any, Dict, List, Tuple
from uuid import uuid4

import numpy as np
import pandas as pd
from fastapi import BackgroundTasks, Body, FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from scipy.stats import f as fisher_f
from scipy.stats import studentized_range
from scipy.stats import t as student_t

app = FastAPI(title="ANOVA + Tukey + LSD Fisher (grouped) API")

ALLOWED_ORIGINS = [
    "*",
    "http://127.0.0.1:5500",
    "http://localhost:5500",
    "http://127.0.0.1:5501",
    "http://localhost:5501",
    "http://127.0.0.1:8080",
    "http://localhost:8080",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

jobs: Dict[str, Dict[str, Any]] = {}
jobs_lock = threading.Lock()

MAX_ROWS = 100000
MAX_GROUPS = 5000
MAX_TREATMENTS_PER_GROUP = 80
JOB_TTL_SECONDS = 60 * 60 * 6


@app.get("/")
def root():
    return FileResponse("index.html")


@app.get("/styles.css")
def styles():
    return FileResponse("styles.css", media_type="text/css")


@app.get("/app.js")
def frontend_js():
    return FileResponse("app.js", media_type="application/javascript")


@app.get("/bayer-logo.jpg")
def bayer_logo():
    return FileResponse("bayer-logo.jpg", media_type="image/jpeg")


def _now_ts() -> float:
    return time.time()


def _touch_job(job_id: str) -> None:
    with jobs_lock:
        if job_id in jobs:
            jobs[job_id]["updated_at"] = _now_ts()


def _set_job(job_id: str, **kwargs) -> None:
    with jobs_lock:
        if job_id in jobs:
            jobs[job_id].update(kwargs)
            jobs[job_id]["updated_at"] = _now_ts()


def _cleanup_old_jobs() -> None:
    cutoff = _now_ts() - JOB_TTL_SECONDS
    to_delete: List[str] = []

    with jobs_lock:
        for job_id, payload in jobs.items():
            if float(payload.get("updated_at", 0)) < cutoff:
                to_delete.append(job_id)

        for job_id in to_delete:
            del jobs[job_id]


def _compact_letter_display(pairs_df: pd.DataFrame, treatments: List[str]) -> Dict[str, str]:
    tset = list(treatments)
    idx = {t: i for i, t in enumerate(tset)}
    n = len(tset)

    nodiff = np.eye(n, dtype=bool)

    for _, r in pairs_df.iterrows():
        g1 = str(r["group1"])
        g2 = str(r["group2"])
        rej = bool(r["reject"])
        if g1 in idx and g2 in idx:
            i, j = idx[g1], idx[g2]
            nodiff[i, j] = not rej
            nodiff[j, i] = not rej

    remaining = set(tset)
    letter_groups: List[Tuple[str, List[str]]] = []
    letters = [chr(c) for c in range(ord("a"), ord("z") + 1)]

    letter_i = 0
    while remaining:
        if letter_i < 26:
            letter = letters[letter_i]
        else:
            prefix = letters[(letter_i // 26) - 1]
            suffix = letters[letter_i % 26]
            letter = prefix + suffix

        rem_list = list(remaining)

        degrees = []
        for t in rem_list:
            i = idx[t]
            deg = sum(nodiff[i, idx[x]] for x in rem_list)
            degrees.append((deg, t))
        degrees.sort(reverse=True)
        seed = degrees[0][1]

        group = [seed]
        for cand in rem_list:
            if cand == seed:
                continue
            ok = True
            for member in group:
                if not nodiff[idx[cand], idx[member]]:
                    ok = False
                    break
            if ok:
                group.append(cand)

        letter_groups.append((letter, group))
        for t in group:
            remaining.discard(t)

        letter_i += 1

    out = {t: "" for t in tset}
    for letter, members in letter_groups:
        for t in tset:
            if all(nodiff[idx[t], idx[m]] for m in members):
                out[t] += letter

    for t in tset:
        if out[t] == "":
            out[t] = "a"

    return out


def _relabel_letters_by_mean(
    summary_df: pd.DataFrame,
    letters_map: Dict[str, str],
) -> Dict[str, str]:
    if summary_df.empty or not letters_map:
        return letters_map

    df = summary_df[["treatment", "mean"]].copy()
    df["treatment"] = df["treatment"].astype(str)
    df = df.sort_values("mean", ascending=False).reset_index(drop=True)

    seen = []
    for _, row in df.iterrows():
        trt = row["treatment"]
        raw_letters = str(letters_map.get(trt, "")).strip().lower()
        for ch in raw_letters:
            if ch not in seen:
                seen.append(ch)

    if not seen:
        return {k: "A" for k in letters_map.keys()}

    new_symbols = [chr(c) for c in range(ord("A"), ord("Z") + 1)]
    remap = {}

    for i, old_sym in enumerate(seen):
        if i < 26:
            remap[old_sym] = new_symbols[i]
        else:
            prefix = new_symbols[(i // 26) - 1]
            suffix = new_symbols[i % 26]
            remap[old_sym] = prefix + suffix

    out = {}
    for trt, raw_letters in letters_map.items():
        raw_letters = str(raw_letters).strip().lower()

        rebuilt = []
        for ch in raw_letters:
            if ch in remap and remap[ch] not in rebuilt:
                rebuilt.append(remap[ch])

        out[str(trt)] = "".join(rebuilt) if rebuilt else "A"

    return out


def _to_numeric_series_strong(s: pd.Series) -> pd.Series:
    s2 = s.astype(str).str.strip()
    s2 = s2.str.replace(r"[^0-9,\.\-]", "", regex=True)

    def _one(x: str):
        x = str(x).strip()
        if x in ("", "-", ".", ","):
            return np.nan

        if "," in x and "." in x:
            if x.rfind(",") > x.rfind("."):
                x = x.replace(".", "").replace(",", ".")
            else:
                x = x.replace(",", "")
        else:
            if "," in x:
                x = x.replace(".", "")
                x = x.replace(",", ".")

        try:
            return float(x)
        except Exception:
            return np.nan

    return s2.apply(_one)


def _reject_between(pairs_df: pd.DataFrame, a: str, b: str) -> bool:
    row = pairs_df[
        ((pairs_df["group1"] == a) & (pairs_df["group2"] == b)) |
        ((pairs_df["group1"] == b) & (pairs_df["group2"] == a))
    ]
    if row.empty:
        return False
    return bool(row.iloc[0]["reject"])


def _make_class_from_pairs(summary_df: pd.DataFrame, pairs_df: pd.DataFrame, class_col_name: str) -> pd.Series:
    df = summary_df.sort_values("mean", ascending=False).reset_index(drop=True)

    if df.empty:
        return pd.Series(dtype=str)

    current = "A"
    classes = [current]

    for i in range(1, len(df)):
        t_prev = str(df.loc[i - 1, "treatment"])
        t_curr = str(df.loc[i, "treatment"])

        if _reject_between(pairs_df, t_prev, t_curr):
            current = chr(ord(current) + 1)

        classes.append(current)

    df[class_col_name] = classes
    return df.set_index("treatment")[class_col_name]


def _run_tukey(
    gdf: pd.DataFrame,
    value_col_num: str,
    trt_col: str,
    alpha: float,
    mse: float,
    df_resid: float,
) -> pd.DataFrame:
    means = (
        gdf.groupby(trt_col, dropna=False)[value_col_num]
        .agg(n="count", mean="mean")
        .reset_index()
        .rename(columns={trt_col: "treatment"})
    )

    treatments = means["treatment"].astype(str).tolist()
    n_map = means.set_index("treatment")["n"].to_dict()
    mean_map = means.set_index("treatment")["mean"].to_dict()
    q_crit = studentized_range.ppf(1 - alpha, len(treatments), df_resid) if df_resid > 0 else np.nan

    rows = []
    for g1, g2 in combinations(treatments, 2):
        mean1 = float(mean_map[g1])
        mean2 = float(mean_map[g2])
        n1 = float(n_map[g1])
        n2 = float(n_map[g2])
        diff = mean2 - mean1
        se = np.sqrt((mse / 2.0) * ((1.0 / n1) + (1.0 / n2))) if n1 > 0 and n2 > 0 else np.nan

        if np.isnan(se) or se == 0 or np.isnan(df_resid) or df_resid <= 0:
            q_stat = np.nan
            p_adj = np.nan
            margin = np.nan
            lower = np.nan
            upper = np.nan
            reject = False
        else:
            q_stat = abs(diff) / se
            p_adj = studentized_range.sf(q_stat, len(treatments), df_resid)
            margin = q_crit * se
            lower = diff - margin
            upper = diff + margin
            reject = bool(p_adj < alpha)

        rows.append({
            "group1": g1,
            "group2": g2,
            "meandiff": diff,
            "p_adj": p_adj,
            "lower": lower,
            "upper": upper,
            "reject": reject,
            "method": "tukey",
            "pvalue": p_adj,
        })

    tuk_df = pd.DataFrame(rows)
    if tuk_df.empty:
        tuk_df = pd.DataFrame(columns=[
            "group1", "group2", "meandiff", "p_adj", "lower", "upper", "reject", "method", "pvalue"
        ])

    return tuk_df


def _run_lsd_fisher(
    gdf: pd.DataFrame,
    value_col_num: str,
    trt_col: str,
    alpha: float,
    mse: float,
    df_resid: float,
    anova_pvalue: float,
) -> pd.DataFrame:
    means = (
        gdf.groupby(trt_col, dropna=False)[value_col_num]
        .agg(n="count", mean="mean")
        .reset_index()
        .rename(columns={trt_col: "treatment"})
    )

    uniq_trt = means["treatment"].astype(str).tolist()
    n_map = means.set_index("treatment")["n"].to_dict()
    mean_map = means.set_index("treatment")["mean"].to_dict()

    rows = []
    for g1, g2 in combinations(uniq_trt, 2):
        mean1 = float(mean_map[g1])
        mean2 = float(mean_map[g2])
        n1 = float(n_map[g1])
        n2 = float(n_map[g2])

        se = np.sqrt(mse * ((1.0 / n1) + (1.0 / n2))) if n1 > 0 and n2 > 0 else np.nan
        diff = mean2 - mean1

        if np.isnan(se) or se == 0 or np.isnan(df_resid) or df_resid <= 0:
            t_stat = np.nan
            p_val = np.nan
            t_crit = np.nan
            lsd_value = np.nan
            lower = np.nan
            upper = np.nan
            reject = False
        else:
            t_stat = abs(diff) / se
            p_val = 2 * (1 - student_t.cdf(abs(t_stat), df_resid))
            t_crit = student_t.ppf(1 - (alpha / 2), df_resid)
            lsd_value = t_crit * se
            lower = diff - lsd_value
            upper = diff + lsd_value
            reject = bool(abs(diff) > lsd_value)

        if np.isnan(anova_pvalue) or anova_pvalue >= alpha:
            reject = False

        rows.append({
            "group1": g1,
            "group2": g2,
            "mean1": mean1,
            "mean2": mean2,
            "meandiff": diff,
            "se": se,
            "t_stat": t_stat,
            "pvalue": p_val,
            "lower": lower,
            "upper": upper,
            "lsd_value": lsd_value,
            "reject": reject,
            "method": "lsd_fisher",
        })

    lsd_df = pd.DataFrame(rows)
    if lsd_df.empty:
        lsd_df = pd.DataFrame(columns=[
            "group1", "group2", "mean1", "mean2", "meandiff", "se", "t_stat",
            "pvalue", "lower", "upper", "lsd_value", "reject", "method"
        ])

    return lsd_df


def _run_group_analysis(
    gdf: pd.DataFrame,
    value_col_num: str,
    trt_col: str,
    alpha: float,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    gdf = gdf.copy()

    gdf = gdf[[value_col_num, trt_col]].dropna()
    if gdf.empty:
        raise ValueError("Grupo sin datos luego de limpiar NA.")

    gdf[value_col_num] = pd.to_numeric(gdf[value_col_num], errors="coerce")
    gdf = gdf.dropna(subset=[value_col_num])
    if gdf.empty:
        raise ValueError(f"Grupo sin valores numéricos en {value_col_num}.")

    gdf[trt_col] = gdf[trt_col].astype(str)
    uniq_trt = sorted(gdf[trt_col].unique().tolist())
    if len(uniq_trt) < 2:
        raise ValueError("Grupo con menos de 2 tratamientos (no se puede ANOVA/post hoc).")

    if len(uniq_trt) > MAX_TREATMENTS_PER_GROUP:
        raise ValueError(
            f"Grupo con demasiados tratamientos ({len(uniq_trt)}). "
            f"Máximo permitido: {MAX_TREATMENTS_PER_GROUP}."
        )

    grouped_values = [
        pd.to_numeric(group[value_col_num], errors="coerce").dropna().astype(float).to_numpy()
        for _, group in gdf.groupby(trt_col, dropna=False, sort=False)
    ]
    grouped_values = [values for values in grouped_values if len(values) > 0]
    k_groups = len(grouped_values)
    n_total = int(sum(len(values) for values in grouped_values))
    if k_groups < 2 or n_total <= k_groups:
        raise ValueError("Grupo sin replicacion suficiente para estimar error residual.")

    grand_mean = float(np.concatenate(grouped_values).mean())
    ss_between = float(sum(len(values) * ((float(values.mean()) - grand_mean) ** 2) for values in grouped_values))
    ss_within = float(sum(((values - float(values.mean())) ** 2).sum() for values in grouped_values))
    df_between = float(k_groups - 1)
    df_resid = float(n_total - k_groups)
    ms_between = ss_between / df_between if df_between > 0 else np.nan
    mse = ss_within / df_resid if df_resid > 0 else np.nan
    f_stat = ms_between / mse if mse and not np.isnan(mse) and mse > 0 else np.nan
    p_anova = float(fisher_f.sf(f_stat, df_between, df_resid)) if not np.isnan(f_stat) else np.nan

    anova_out = pd.DataFrame([{
        "df": df_between,
        "F": f_stat,
        "pvalue": p_anova,
        "df_resid": df_resid,
    }])

    summary = (
        gdf.groupby(trt_col, dropna=False)[value_col_num]
        .agg(n="count", mean="mean", sd="std")
        .reset_index()
        .rename(columns={trt_col: "treatment"})
    )

    tukey_pairs_df = _run_tukey(gdf, value_col_num, trt_col, alpha, mse, df_resid)
    tukey_letters = _compact_letter_display(tukey_pairs_df, uniq_trt)
    tukey_letters = _relabel_letters_by_mean(summary[["treatment", "mean"]].copy(), tukey_letters)

    summary["tukey_letters"] = (
        summary["treatment"].astype(str).map(tukey_letters).fillna("A").str.upper()
    )

    tukey_class_map = _make_class_from_pairs(
        summary[["treatment", "mean"]].copy(),
        tukey_pairs_df,
        "tukey_class",
    )
    summary["tukey_class"] = (
        summary["treatment"].astype(str).map(tukey_class_map).fillna("A").str.upper()
    )

    lsd_pairs_df = _run_lsd_fisher(
        gdf=gdf,
        value_col_num=value_col_num,
        trt_col=trt_col,
        alpha=alpha,
        mse=mse,
        df_resid=df_resid,
        anova_pvalue=p_anova,
    )

    lsd_letters = _compact_letter_display(lsd_pairs_df, uniq_trt) if not lsd_pairs_df.empty else {t: "a" for t in uniq_trt}
    lsd_letters = _relabel_letters_by_mean(summary[["treatment", "mean"]].copy(), lsd_letters)

    summary["lsd_letters"] = (
        summary["treatment"].astype(str).map(lsd_letters).fillna("A").str.upper()
    )

    lsd_class_map = (
        _make_class_from_pairs(
            summary[["treatment", "mean"]].copy(),
            lsd_pairs_df,
            "lsd_class",
        )
        if not lsd_pairs_df.empty
        else pd.Series({t: "A" for t in uniq_trt})
    )
    summary["lsd_class"] = (
        summary["treatment"].astype(str).map(lsd_class_map).fillna("A").str.upper()
    )

    pairs_df = pd.concat([tukey_pairs_df, lsd_pairs_df], ignore_index=True, sort=False)

    return summary, anova_out, pairs_df



def _norm_text(value: Any) -> str:
    return str(value).strip().lower()


def _norm_key(value: Any) -> str:
    return str(value).strip()


def _is_excluded_control_row(
    row: pd.Series,
    se_name_mod_col: str,
    treatment_col: str,
    control_rules: Dict[str, Dict[str, Any]],
) -> bool:
    if not se_name_mod_col or se_name_mod_col not in row.index:
        return False

    se_value = _norm_key(row.get(se_name_mod_col, ""))
    rule = control_rules.get(se_value) or control_rules.get(_norm_text(se_value))
    if not rule:
        return False

    include_control = bool(rule.get("include_control", True))
    control_treatment = _norm_key(rule.get("control_treatment", "1") or "1")

    if include_control:
        return False

    return _norm_key(row.get(treatment_col, "")) == control_treatment


def _placeholder_summary_rows_for_excluded_controls(
    df: pd.DataFrame,
    treatment_col: str,
    group_cols: List[str],
) -> pd.DataFrame:
    if "excluded_from_stats" not in df.columns:
        return pd.DataFrame()

    excluded = df[df["excluded_from_stats"] == True].copy()
    if excluded.empty:
        return pd.DataFrame()

    rows = []
    cols = (group_cols if group_cols else []) + [treatment_col]
    for _, r in excluded[cols].drop_duplicates().iterrows():
        item = {"treatment": str(r[treatment_col])}
        for c in group_cols:
            item[c] = r[c]
        item.update({
            "n": "-",
            "mean": "-",
            "sd": "-",
            "tukey_letters": "-",
            "tukey_class": "-",
            "lsd_letters": "-",
            "lsd_class": "-",
            "stats_status": "excluded_control",
        })
        rows.append(item)

    return pd.DataFrame(rows)

def _make_group_key(row: pd.Series, group_cols: List[str]) -> str:
    parts = []
    for c in group_cols:
        v = row.get(c, "")
        if pd.isna(v):
            v = ""
        parts.append(f"{c}={v}")
    return " | ".join(parts)



def _style_excel_workbook(writer) -> None:
    """Aplica un formato corporativo Bayer al Excel exportado."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="00B5E2")  # celeste Bayer
    sub_fill = PatternFill("solid", fgColor="EAF8FC")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    thick_side = Side(style="medium", color="111111")
    thin_side = Side(style="thin", color="D9D9D9")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    for ws in writer.book.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        if ws.max_row >= 1:
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = Border(left=thick_side, right=thick_side, top=thick_side, bottom=thick_side)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.fill = white_fill
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=False)
        # Resalta columnas de alcance del análisis
        for col_idx, cell in enumerate(ws[1], start=1):
            if str(cell.value or "") in {"analysis_scope", "analysis_basis", "location_analysis_note", "group_key"}:
                for r in range(1, ws.max_row + 1):
                    ws.cell(r, col_idx).fill = header_fill if r == 1 else sub_fill
                    if r == 1:
                        ws.cell(r, col_idx).font = Font(bold=True, color="FFFFFF")
        for col_idx in range(1, ws.max_column + 1):
            letter = get_column_letter(col_idx)
            max_len = 10
            for cell in ws[letter]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, min(len(value), 45))
            ws.column_dimensions[letter].width = min(max(max_len + 2, 11), 38)
        ws.auto_filter.ref = ws.dimensions


def _merge_scope_results(
    df_scope: pd.DataFrame,
    summary_df: pd.DataFrame,
    anova_df: pd.DataFrame,
    treatment_col: str,
    group_cols: List[str],
) -> pd.DataFrame:
    base_df = df_scope.copy()
    left_merge_keys = ["analysis_scope"] + (group_cols if group_cols else []) + [treatment_col]
    right_merge_keys = ["analysis_scope"] + (group_cols if group_cols else []) + ["treatment"]

    if not summary_df.empty:
        final_df = base_df.merge(
            summary_df,
            left_on=left_merge_keys,
            right_on=right_merge_keys,
            how="left",
        )
        if treatment_col != "treatment" and "treatment" in final_df.columns:
            final_df = final_df.drop(columns=["treatment"])
    else:
        final_df = base_df

    if not anova_df.empty:
        anova_cols = [c for c in ["df", "F", "pvalue", "df_resid", "error"] if c in anova_df.columns]
        key_cols = ["analysis_scope"] + (group_cols if group_cols else [])
        available_key_cols = [c for c in key_cols if c in anova_df.columns]
        a_small = anova_df[available_key_cols + anova_cols].copy().drop_duplicates()
        if available_key_cols:
            final_df = final_df.merge(a_small, on=available_key_cols, how="left", suffixes=("", "_anova"))
        elif len(anova_df):
            for c in anova_cols:
                final_df[c] = anova_df.iloc[0][c]

    stats_cols = [
        "n", "mean", "sd", "tukey_letters", "tukey_class", "lsd_letters", "lsd_class",
        "df", "F", "pvalue", "df_resid"
    ]
    if "excluded_from_stats" in final_df.columns:
        mask = final_df["excluded_from_stats"] == True
        for c in stats_cols:
            if c in final_df.columns:
                final_df[c] = final_df[c].astype(object)
                final_df.loc[mask, c] = "-"
        if "stats_status" not in final_df.columns:
            final_df["stats_status"] = "analyzed"
        final_df["stats_status"] = final_df["stats_status"].fillna("analyzed")
        final_df.loc[mask, "stats_status"] = "excluded_control"

    return final_df


def _analyze_scope(
    job_id: str,
    df: pd.DataFrame,
    treatment_col: str,
    group_cols: List[str],
    alpha: float,
    scope_label: str,
    scope_note: str,
    group_offset: int,
    total_groups_all: int,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, int]:
    df_scope = df.copy()
    df_scope["analysis_scope"] = scope_label
    df_scope["analysis_basis"] = "por_localidad" if scope_label == "Por localidad" else "por_protocolo"
    df_scope["location_analysis_note"] = scope_note
    df_scope["group_key"] = df_scope.apply(lambda r: _make_group_key(r, group_cols), axis=1) if group_cols else "ALL"

    summaries: List[pd.DataFrame] = []
    anovas: List[pd.DataFrame] = []
    pairs: List[pd.DataFrame] = []

    value_col_num = "assessment_value_num"
    analysis_df = df_scope[df_scope["excluded_from_stats"] != True].copy()
    if analysis_df.empty:
        raise ValueError("No quedaron filas analizables luego de excluir testigos según se_name_mod.")

    grouped_items = list(analysis_df.groupby(group_cols, dropna=False, sort=False)) if group_cols else [("ALL", analysis_df)]
    if len(grouped_items) > MAX_GROUPS:
        raise ValueError(f"Demasiados grupos ({len(grouped_items)}). Máximo permitido: {MAX_GROUPS}.")

    for idx_group, (keys, gdf) in enumerate(grouped_items, start=1):
        key_dict: Dict[str, Any] = {"analysis_scope": scope_label}
        if group_cols:
            if isinstance(keys, tuple):
                for col, val in zip(group_cols, keys):
                    key_dict[col] = val
            else:
                key_dict[group_cols[0]] = keys
        try:
            s, a, p = _run_group_analysis(gdf, value_col_num, treatment_col, alpha)
            for col, val in key_dict.items():
                s[col] = val
                a[col] = val
                p[col] = val
            summaries.append(s)
            anovas.append(a)
            pairs.append(p)
        except Exception as e:
            err = {"error": str(e), "analysis_scope": scope_label}
            for col, val in key_dict.items():
                err[col] = val
            anovas.append(pd.DataFrame([err]))

        done = group_offset + idx_group
        progress = int(5 + (done / max(total_groups_all, 1)) * 88)
        _set_job(
            job_id,
            current=done,
            total=total_groups_all,
            progress=min(progress, 95),
            message=f"Procesando {done}/{total_groups_all} grupos..."
        )

    summary_df = pd.concat(summaries, ignore_index=True) if summaries else pd.DataFrame()
    excluded_summary_df = _placeholder_summary_rows_for_excluded_controls(df_scope, treatment_col, group_cols)
    if not excluded_summary_df.empty:
        excluded_summary_df["analysis_scope"] = scope_label
        summary_df = pd.concat([summary_df, excluded_summary_df], ignore_index=True, sort=False)
    anova_df = pd.concat(anovas, ignore_index=True) if anovas else pd.DataFrame()
    pairs_df = pd.concat(pairs, ignore_index=True) if pairs else pd.DataFrame()
    result_df = _merge_scope_results(df_scope, summary_df, anova_df, treatment_col, group_cols)
    return result_df, summary_df, anova_df, pairs_df, len(grouped_items)


def _build_excel_output(
    final_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    anova_df: pd.DataFrame,
    pairs_df: pd.DataFrame,
    analysis_name: str,
    original_columns: List[str],
) -> io.BytesIO:
    output = io.BytesIO()

    original_front = [col for col in original_columns if col in final_df.columns]
    calculated_cols = [col for col in final_df.columns if col not in original_front]
    final_df = final_df[original_front + calculated_cols]

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        final_df.to_excel(writer, index=False, sheet_name="results")
        if not summary_df.empty:
            summary_df.to_excel(writer, index=False, sheet_name="summary_by_treatment")
        if not anova_df.empty:
            anova_df.to_excel(writer, index=False, sheet_name="anova_detail")
        if not pairs_df.empty:
            tukey_df = pairs_df[pairs_df["method"] == "tukey"].copy() if "method" in pairs_df.columns else pd.DataFrame()
            lsd_df = pairs_df[pairs_df["method"] == "lsd_fisher"].copy() if "method" in pairs_df.columns else pd.DataFrame()
            if not tukey_df.empty:
                tukey_df.to_excel(writer, index=False, sheet_name="tukey_pairs_detail")
            if not lsd_df.empty:
                lsd_df.to_excel(writer, index=False, sheet_name="lsd_pairs_detail")
        scope_readme = pd.DataFrame([
            {"campo": "Por localidad", "significado": "El análisis se calcula separando cada localidad. La localidad forma parte del group_key."},
            {"campo": "Por protocolo", "significado": "El análisis se calcula juntando todas las localidades. La localidad NO forma parte del modelo/grupo estadístico."},
            {"campo": "analysis_scope", "significado": "Indica si esa fila corresponde al análisis por localidad o al análisis por protocolo."},
            {"campo": "location_analysis_note", "significado": "Aclara si la línea fue analizada discriminando localidad o juntando localidades."},
        ])
        scope_readme.to_excel(writer, index=False, sheet_name="analysis_scope_readme")
        _style_excel_workbook(writer)

    output.seek(0)
    return output


def _unique_keep_order(items: List[str]) -> List[str]:
    out: List[str] = []
    for x in items:
        if x and x not in out:
            out.append(x)
    return out


def _run_analysis_job(job_id: str, payload: Dict[str, Any]) -> None:
    try:
        rows = payload.get("rows")
        value_col = payload.get("value_col", "assessment_value")
        treatment_col = payload.get("treatment_col", "treatment")
        group_cols = payload.get("group_cols", []) or []
        alpha = float(payload.get("alpha", 0.05))
        analysis_name = str(payload.get("analysis_name", "")).strip()
        analysis_scope = str(payload.get("analysis_scope", "location")).strip().lower()
        location_col = str(payload.get("location_col", "")).strip()
        se_name_mod_col = str(payload.get("se_name_mod_col", "se_name_mod")).strip()
        control_rules_raw = payload.get("se_name_mod_control_rules", {}) or {}
        control_rules: Dict[str, Dict[str, Any]] = {}
        if isinstance(control_rules_raw, dict):
            for key, rule in control_rules_raw.items():
                if isinstance(rule, dict):
                    control_rules[_norm_key(key)] = rule
                    control_rules[_norm_text(key)] = rule

        if not isinstance(rows, list) or len(rows) == 0:
            raise ValueError("rows vacío o inválido.")
        if analysis_name == "":
            raise ValueError("analysis_name es requerido.")
        if len(rows) > MAX_ROWS:
            raise ValueError(f"Dataset demasiado grande. Máximo permitido: {MAX_ROWS} filas.")
        if analysis_scope not in {"location", "protocol", "both"}:
            analysis_scope = "location"

        df = pd.DataFrame(rows)
        original_columns = list(df.columns)
        missing = [c for c in [value_col, treatment_col] if c not in df.columns]
        if missing:
            raise ValueError(f"Faltan columnas requeridas: {missing}")
        for c in group_cols:
            if c not in df.columns:
                raise ValueError(f"Columna de agrupamiento no existe: {c}")
        if location_col and location_col not in df.columns:
            raise ValueError(f"La columna de localidad seleccionada no existe: {location_col}")
        if analysis_scope in {"location", "both"} and not location_col:
            raise ValueError("Para analizar por localidad o ambas opciones, seleccioná una columna de localidad.")
        if se_name_mod_col and se_name_mod_col not in df.columns:
            se_name_mod_col = ""

        df[treatment_col] = df[treatment_col].astype(str)
        if se_name_mod_col and control_rules:
            df["excluded_from_stats"] = df.apply(
                lambda r: _is_excluded_control_row(r, se_name_mod_col, treatment_col, control_rules),
                axis=1,
            )
        else:
            df["excluded_from_stats"] = False

        _set_job(job_id, progress=2, message="Convirtiendo valores a numéricos...")
        df["assessment_value_num"] = _to_numeric_series_strong(df[value_col])
        df["assessment_value_x1"] = df["assessment_value_num"] * 1.0
        df = df.dropna(subset=["assessment_value_num"])
        if df.empty:
            raise ValueError(f"No quedaron filas con valores numéricos en '{value_col}'.")
        df["analysis_name"] = analysis_name

        base_group_cols = _unique_keep_order([c for c in group_cols if c not in {value_col, treatment_col}])
        scopes: List[Tuple[str, List[str], str]] = []
        if analysis_scope in {"location", "both"}:
            local_cols = _unique_keep_order(base_group_cols + ([location_col] if location_col else []))
            scopes.append(("Por localidad", local_cols, "Línea analizada por localidad: la localidad entra como corte del análisis."))
        if analysis_scope in {"protocol", "both"}:
            protocol_cols = _unique_keep_order([c for c in base_group_cols if c != location_col])
            scopes.append(("Por protocolo", protocol_cols, "Línea analizada por protocolo: se juntan todas las localidades."))

        # Cuenta grupos totales antes de correr para que la barra de progreso sea clara.
        analysis_df_tmp = df[df["excluded_from_stats"] != True].copy()
        total_groups_all = 0
        for _, cols_scope, _ in scopes:
            total_groups_all += len(list(analysis_df_tmp.groupby(cols_scope, dropna=False, sort=False))) if cols_scope else 1
        if total_groups_all > MAX_GROUPS:
            raise ValueError(f"Demasiados grupos totales ({total_groups_all}). Máximo permitido: {MAX_GROUPS}.")
        _set_job(job_id, total=total_groups_all, current=0, progress=5, status="running", message=f"Procesando 0/{total_groups_all} grupos...")

        all_results: List[pd.DataFrame] = []
        all_summaries: List[pd.DataFrame] = []
        all_anovas: List[pd.DataFrame] = []
        all_pairs: List[pd.DataFrame] = []
        done_offset = 0
        for scope_label, cols_scope, note in scopes:
            result_df, summary_df, anova_df, pairs_df, groups_done = _analyze_scope(
                job_id=job_id,
                df=df,
                treatment_col=treatment_col,
                group_cols=cols_scope,
                alpha=alpha,
                scope_label=scope_label,
                scope_note=note,
                group_offset=done_offset,
                total_groups_all=total_groups_all,
            )
            done_offset += groups_done
            all_results.append(result_df)
            all_summaries.append(summary_df)
            all_anovas.append(anova_df)
            all_pairs.append(pairs_df)

        _set_job(job_id, progress=96, message="Armando resultados...")
        final_df = pd.concat(all_results, ignore_index=True, sort=False) if all_results else pd.DataFrame()
        summary_df = pd.concat(all_summaries, ignore_index=True, sort=False) if all_summaries else pd.DataFrame()
        anova_df = pd.concat(all_anovas, ignore_index=True, sort=False) if all_anovas else pd.DataFrame()
        pairs_df = pd.concat(all_pairs, ignore_index=True, sort=False) if all_pairs else pd.DataFrame()

        output = _build_excel_output(
            final_df=final_df,
            summary_df=summary_df,
            anova_df=anova_df,
            pairs_df=pairs_df,
            analysis_name=analysis_name,
            original_columns=original_columns,
        )
        safe_name = "".join(ch if ch.isalnum() or ch in (" ", "_", "-") else "_" for ch in analysis_name).strip() or "analysis"
        suffix = {"location": "por_localidad", "protocol": "por_protocolo", "both": "ambos"}.get(analysis_scope, "analisis")
        filename = f"{safe_name}_anova_tukey_lsd_{suffix}.xlsx"
        _set_job(job_id, status="done", progress=100, message="Análisis finalizado.", result_bytes=output.getvalue(), filename=filename)
    except Exception as e:
        _set_job(job_id, status="error", progress=100, message="El análisis terminó con error.", error=str(e))

@app.post("/analyze")
def analyze(
    background_tasks: BackgroundTasks,
    payload: Dict[str, Any] = Body(...),
):
    _cleanup_old_jobs()

    job_id = str(uuid4())

    with jobs_lock:
        jobs[job_id] = {
            "job_id": job_id,
            "status": "running",
            "progress": 0,
            "current": 0,
            "total": 0,
            "message": "Job creado.",
            "error": None,
            "result_bytes": None,
            "filename": None,
            "created_at": _now_ts(),
            "updated_at": _now_ts(),
        }

    background_tasks.add_task(_run_analysis_job, job_id, payload)

    return {"job_id": job_id}


@app.get("/status/{job_id}")
def status(job_id: str):
    with jobs_lock:
        job = jobs.get(job_id)

        if not job:
            raise HTTPException(status_code=404, detail="Job no encontrado")

        return {
            "job_id": job["job_id"],
            "status": job["status"],
            "progress": job["progress"],
            "current": job["current"],
            "total": job["total"],
            "message": job.get("message"),
            "error": job.get("error"),
            "filename": job.get("filename"),
        }


@app.get("/download/{job_id}")
def download(job_id: str):
    with jobs_lock:
        job = jobs.get(job_id)

        if not job:
            raise HTTPException(status_code=404, detail="Job no encontrado")

        if job["status"] != "done":
            raise HTTPException(status_code=400, detail="El archivo todavía no está listo.")

        result_bytes = job.get("result_bytes")
        filename = job.get("filename") or "analysis_anova_tukey_lsd.xlsx"

    if result_bytes is None:
        raise HTTPException(status_code=500, detail="No se encontró el archivo generado.")

    output = io.BytesIO(result_bytes)
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/health")
def health():
    return {"status": "ok"}


@app.get("/version")
def version():
    return {"version": "2026-06-02-location-scope-both-v1"}
